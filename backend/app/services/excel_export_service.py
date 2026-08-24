"""
Excel Export Service.

Builds the tracker's Excel export — filterable (date range, location,
recruiter, Account Manager, status), per the design decision that this
is for periodic incentive/reporting use, not a single unfiltered dump
someone then has to filter themselves in Excel afterward.

Recruiter and Account Manager NAMES (not just their IDs) are resolved
and included as columns, since the whole point of this export is
incentive calculation — a sheet full of UUIDs would be useless for that.
"""

from __future__ import annotations

from io import BytesIO
from uuid import UUID

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from app.db.client import get_service_db
from app.models.user import Candidate

_COLUMNS = [
    ("employee_id", "Employee ID"),
    ("full_name", "Candidate Name"),
    ("email", "Email"),
    ("phone", "Phone"),
    ("client_name", "Client"),
    ("designation", "Designation"),
    ("department", "Department"),
    ("account_manager_name", "Account Manager"),
    ("recruiter_name", "Recruiter"),
    ("location_name", "Location"),
    ("stage", "Stage"),
    ("proposed_ctc", "Proposed CTC"),
    ("expected_doj", "Expected DOJ"),
    ("confirmed_doj", "Confirmed DOJ"),
    ("offer_released_at", "Offer Released"),
    ("appointment_released_at", "Appointment Released"),
    ("last_working_day", "Last Working Day"),
    ("relieving_released_at", "Relieving Released"),
    ("request_date", "Request Date"),
]


class ExcelExportService:
    def __init__(self) -> None:
        self._db = get_service_db()

    def _resolve_names(self, candidates: list[Candidate]) -> dict[str, dict[str, str]]:
        """
        One batch lookup per lookup table rather than one query per
        candidate — an export covering hundreds of candidates shouldn't
        mean hundreds of extra round trips just to turn IDs into names.
        """
        am_ids = {str(c.account_manager_id) for c in candidates if c.account_manager_id}
        recruiter_ids = {str(c.recruiter_id) for c in candidates if c.recruiter_id}
        location_ids = {str(c.location_id) for c in candidates}

        am_names: dict[str, str] = {}
        if am_ids:
            result = self._db.table("directory_account_managers").select("id, full_name").in_("id", list(am_ids)).execute()
            am_names = {row["id"]: row["full_name"] for row in result.data}

        recruiter_names: dict[str, str] = {}
        if recruiter_ids:
            result = self._db.table("directory_recruiters").select("id, full_name").in_("id", list(recruiter_ids)).execute()
            recruiter_names = {row["id"]: row["full_name"] for row in result.data}

        location_names: dict[str, str] = {}
        if location_ids:
            result = self._db.table("locations").select("id, name").in_("id", list(location_ids)).execute()
            location_names = {row["id"]: row["name"] for row in result.data}

        return {"am": am_names, "recruiter": recruiter_names, "location": location_names}

    def build_export(self, candidates: list[Candidate]) -> bytes:
        names = self._resolve_names(candidates)

        wb = Workbook()
        ws = wb.active
        ws.title = "Candidates"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="7C3AED", end_color="7C3AED", fill_type="solid")

        for col_index, (_, label) in enumerate(_COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_index, value=label)
            cell.font = header_font
            cell.fill = header_fill

        for row_index, candidate in enumerate(candidates, start=2):
            row_data = {
                "employee_id": candidate.employee_id or "",
                "full_name": candidate.full_name,
                "email": candidate.email,
                "phone": candidate.phone or "",
                "client_name": candidate.client_name,
                "designation": candidate.designation or "",
                "department": candidate.department or "",
                "account_manager_name": names["am"].get(str(candidate.account_manager_id), ""),
                "recruiter_name": names["recruiter"].get(str(candidate.recruiter_id), ""),
                "location_name": names["location"].get(str(candidate.location_id), ""),
                "stage": candidate.stage,
                "proposed_ctc": candidate.proposed_ctc or "",
                "expected_doj": str(candidate.expected_doj) if candidate.expected_doj else "",
                "confirmed_doj": str(candidate.confirmed_doj) if candidate.confirmed_doj else "",
                "offer_released_at": str(candidate.offer_released_at) if candidate.offer_released_at else "",
                "appointment_released_at": str(candidate.appointment_released_at) if candidate.appointment_released_at else "",
                "last_working_day": str(candidate.last_working_day) if candidate.last_working_day else "",
                "relieving_released_at": str(candidate.relieving_released_at) if candidate.relieving_released_at else "",
                "request_date": str(candidate.request_date),
            }
            for col_index, (field_key, _) in enumerate(_COLUMNS, start=1):
                ws.cell(row=row_index, column=col_index, value=row_data.get(field_key, ""))

        for col_index, (_, label) in enumerate(_COLUMNS, start=1):
            ws.column_dimensions[get_column_letter(col_index)].width = max(len(label) + 2, 14)

        ws.freeze_panes = "A2"

        buffer = BytesIO()
        wb.save(buffer)
        return buffer.getvalue()
